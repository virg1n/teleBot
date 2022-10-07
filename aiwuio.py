import telebot
import random
import openpyxl
from telebot import types
ans = ""
letters, correctwords, players, points, baza, counter_addmes, check_true = set(), list(), set(), dict(), dict(), 0, 0
book = openpyxl.load_workbook(filename="da.xlsx")
words = ['год', 'человек', 'время', 'дело', 'жизнь', 'день', 'рука', 'работа', 'слово', 'место', 'вопрос', 'лицо', 'глаз', 'страна', 'друг', 'сторона', 'дом', 'случай', 'ребенок', 'голова', 'система', 'вид', 'конец', 'отношение', 'город', 'часть', 'женщина', 'проблема', 'земля', 'решение', 'власть', 'машина', 'закон', 'час', 'образ', 'отец', 'история', 'нога', 'вода', 'война', 'возможность', 'компания', 'результат', 'дверь', 'бог', 'народ', 'область', 'число', 'голос', 'развитие', 'группа', 'жена', 'процесс', 'условие', 'книга', 'ночь', 'суд', 'деньга', 'уровень', 'начало', 'государство', 'стол', 'средство', 'связь', 'имя', 'президент', 'форма', 'путь', 'организация', 'качество', 'действие', 'статья', 'общество', 'ситуация', 'деятельность', 'школа', 'душа', 'дорога', 'язык', 'взгляд', 'момент', 'минута', 'месяц', 'порядок', 'цель', 'программа', 'муж', 'помощь', 'мысль', 'вечер', 'орган', 'правительство', 'рынок', 'предприятие', 'партия', 'роль', 'смысл', 'мама', 'мера', 'улица', 'состояние', 'задача', 'информация', 'театр', 'внимание', 'производство', 'квартира', 'труд', 'тело', 'письмо', 'центр', 'утро', 'мать', 'комната', 'семья', 'сын', 'смерть', 'положение', 'интерес', 'федерация', 'век', 'идея', 'управление', 'автор', 'окно', 'ответ', 'совет', 'разговор', 'мужчина', 'ряд', 'счет', 'мнение', 'цена', 'точка', 'план', 'проект', 'глава', 'материал', 'основа', 'причина', 'движение', 'культура', 'сердце', 'рубль', 'наука', 'документ', 'неделя', 'вещь', 'чувство', 'правило', 'служба', 'газета', 'срок', 'институт', 'член', 'ход', 'стена', 'директор', 'плечо', 'опыт', 'встреча', 'принцип', 'событие', 'структура', 'количество', 'товарищ', 'создание', 'значение', 'объект', 'гражданин', 'очередь', 'период', 'образование', 'состав', 'пример', 'лес', 'исследование', 'девушка', 'данные', 'палец', 'судьба', 'тип', 'метод', 'политика', 'армия', 'брат', 'представитель', 'борьба', 'использование', 'шаг', 'игра', 'участие', 'территория', 'край', 'размер', 'номер', 'район', 'население', 'банк', 'начальник', 'класс', 'зал', 'изменение', 'большинство', 'характер', 'кровь', 'направление', 'позиция', 'герой', 'течение', 'девочка', 'искусство', 'гость', 'воздух', 'мальчик', 'фильм', 'договор', 'регион', 'выбор', 'свобода', 'врач', 'экономика', 'небо', 'факт', 'церковь', 'завод', 'фирма', 'бизнес', 'союз', 'деньги', 'специалист', 'род', 'команда', 'руководитель', 'спина', 'дух', 'музыка', 'способ', 'хозяин', 'поле', 'доллар', 'память', 'природа', 'дерево', 'оценка', 'объем', 'картина', 'процент', 'требование', 'писатель', 'сцена', 'анализ', 'основание', 'повод', 'вариант', 'берег', 'модель', 'степень', 'самолет', 'телефон', 'граница', 'песня', 'половина', 'министр', 'угол', 'зрение', 'предмет', 'литература', 'операция', 'двор', 'спектакль', 'руководство', 'солнце', 'автомобиль', 'родитель', 'участник', 'журнал', 'база', 'пространство', 'защита', 'название', 'стих', 'ум', 'море', 'удар', 'знание', 'солдат', 'миллион', 'строительство', 'технология', 'председатель', 'сон', 'сознание', 'бумага', 'реформа', 'оружие', 'линия', 'текст', 'выход', 'ребята', 'магазин', 'соответствие', 'участок', 'услуга', 'поэт', 'предложение', 'желание', 'пара', 'успех', 'среда', 'возраст', 'комплекс', 'бюджет', 'представление', 'площадь', 'генерал', 'господин', 'дочь', 'понятие', 'кабинет', 'безопасность', 'фонд', 'сфера', 'папа', 'сотрудник', 'продукция', 'будущее', 'продукт', 'содержание', 'художник', 'республика', 'сумма', 'контроль', 'парень', 'ветер', 'хозяйство', 'помочь', 'курс', 'губа', 'река', 'грудь', 'огонь', 'нос', 'волос', 'ухо', 'отсутствие', 'радость', 'сад', 'подготовка', 'необходимость', 'доктор', 'лето', 'камень', 'здание', 'капитан', 'собака', 'итог', 'рис', 'техника', 'элемент', 'источник', 'деревня', 'депутат', 'проведение', 'рот', 'масса', 'комиссия', 'цвет', 'рассказ', 'функция', 'определение', 'мужик', 'обеспечение', 'обстоятельство', 'работник', 'разработка', 'лист', 'звезда', 'гора', 'применение', 'победа', 'товар', 'воля', 'зона', 'предел', 'целое', 'личность', 'офицер', 'влияние', 'поддержка', 'ответственность', 'цветок', 'праздник', 'немец', 'бой', 'сестра', 'господь', 'учитель', 'многое', 'рамка', 'практика', 'показатель', 'метр', 'войско', 'частность', 'особенность', 'снег', 'комитет', 'налог', 'акт', 'отдел', 'карман', 'вывод', 'норма', 'читатель', 'этап', 'сравнение', 'прошлое', 'фамилия', 'кухня', 'заявление', 'доля', 'пункт', 'студент', 'учет', 'впечатление', 'доход', 'вирус', 'клетка', 'удовольствие', 'теория', 'враг', 'собрание', 'бутылка', 'расчет', 'го', 'режим', 'множество', 'клуб', 'попытка', 'зуб', 'сеть', 'семь', 'министерство', 'прием', 'боль', 'сожаление', 'кожа', 'субъект', 'знак', 'актер', 'ресурс', 'акция', 'газ', 'журналист', 'звук', 'передача', 'здоровье', 'администрация', 'болезнь', 'детство', 'мастер', 'выборы', 'зима', 'подход', 'поиск', 'механизм', 'выражение', 'скорость', 'ощущение', 'стоимость', 'коридор', 'ошибка', 'лидер', 'карта', 'заседание', 'глубина', 'хлеб', 'поверхность', 'энергия', 'нарушение', 'реализация', 'революция', 'поведение', 'профессор', 'исполнение', 'заместитель', 'суть', 'станция', 'реакция', 'десяток', 'столица', 'формирование', 'поколение', 'дума', 'существование', 'продажа', 'список', 'способность', 'противник', 'схема', 'долг', 'режиссер', 'отличие', 'колено', 'дед', 'свойство', 'этаж', 'секунда', 'фактор', 'житель', 'явление', 'высота', 'сосед', 'усилие', 'рождение', 'расход', 'остров', 'фигура', 'наличие', 'дядя', 'милиция', 'растение', 'существо', 'черт', 'бабушка', 'законодательство', 'собственность', 'отрасль', 'слеза', 'волна', 'стекло', 'традиция', 'январь', 'оборудование', 'зависимость', 'фраза', 'декабрь', 'сведение', 'трубка', 'сентябрь', 'университет', 'командир', 'храм', 'повышение', 'стиль', 'артист', 'больница', 'одежда', 'охрана', 'водка', 'кодекс', 'имущество', 'птица', 'переход', 'красота', 'клиент', 'толпа', 'адрес', 'отделение', 'октябрь', 'чудо', 'счастие', 'улыбка', 'ужас', 'аппарат', 'корабль', 'родина', 'животное', 'черта', 'известие', 'понимание', 'тень', 'апрель', 'коллега', 'преступление', 'рыба', 'кресло', 'запах', 'выставка', 'князь', 'фотография', 'весна', 'помещение', 'эпоха', 'занятие', 'произведение', 'концерт', 'ладонь', 'дама', 'сомнение', 'американец', 'середина', 'зарплата', 'тайна', 'запад', 'июнь', 'беседа', 'фронт', 'поезд', 'должность', 'баба', 'промышленность', 'музей', 'судья', 'получение', 'полковник', 'зритель', 'секретарь', 'установка', 'поток', 'ценность', 'образец', 'страница', 'перспектива', 'трава', 'чиновник', 'мозг', 'сотня', 'лагерь', 'выступление', 'оборона', 'постановление', 'честь', 'настроение', 'кровать', 'характеристика', 'обязанность', 'шея', 'крыша', 'появление', 'учреждение', 'признак', 'труба', 'жертва', 'беда', 'фон', 'организм', 'ученик', 'заключение', 'выполнение', 'канал', 'исключение', 'дача', 'соглашение', 'осень', 'польза', 'стул', 'июль', 'дождь', 'сутки', 'еврей', 'конкурс', 'открытие', 'телевизор', 'лошадь', 'температура', 'приказ', 'лестница', 'реклама', 'спор', 'подруга', 'угроза', 'конфликт', 'изучение', 'вино', 'концепция', 'достижение', 'сообщение', 'объединение', 'обстановка', 'костюм', 'ключ', 'ресторан', 'назначение', 'царь', 'воспоминание', 'увеличение', 'вкус', 'мероприятие', 'лоб', 'слой', 'восток', 'последствие', 'принятие', 'сотрудничество', 'нефть', 'слух', 'бок', 'переговоры', 'тюрьма', 'кандидат', 'просьба', 'реальность', 'подарок', 'категория', 'потребность', 'быль', 'редакция', 'очко', 'километр', 'губернатор', 'новость', 'инструмент', 'потеря', 'взаимодействие', 'звонок', 'кусок', 'капитал', 'грех', 'перевод', 'партнер', 'ноябрь', 'молодежь', 'тишина', 'творчество', 'книжка', 'мясо', 'масло', 'деталь', 'инженер', 'оплата', 'эксперт', 'кремль', 'февраль', 'следствие', 'пьеса', 'билет', 'урок', 'коллектив', 'устройство', 'палата', 'площадка', 'опасность', 'пропасть', 'воздействие', 'разница', 'родственник', 'сезон', 'издание', 'человечество', 'снижение', 'запас', 'крик', 'публика', 'вещество', 'экран', 'эффект', 'ящик', 'ракета', 'водитель', 'пакет', 'зеркало', 'вес', 'дно', 'вагон', 'убийство', 'тон', 'щека', 'дурак', 'длина', 'давление', 'двигатель', 'камера', 'обращение', 'формула', 'запись', 'крыло', 'поездка', 'гостиница', 'колесо', 'разрешение', 'торговля', 'академия', 'доклад', 'общение', 'присутствие', 'процедура', 'испытание', 'нож', 'проверка', 'коммунист', 'цифра', 'полет', 'стакан', 'эффективность', 'обучение', 'портрет', 'достоинство', 'рассмотрение', 'владелец', 'жилье', 'компьютер', 'корень', 'смена', 'доказательство', 'кадр', 'лейтенант', 'признание', 'темнота', 'пистолет', 'наблюдение', 'мост', 'ремонт', 'истина', 'вход', 'политик', 'живот', 'кредит', 'шум', 'обед', 'недостаток', 'памятник', 'вершина', 'серия', 'эксперимент', 'сущность', 'транспорт', 'инициатива', 'активность', 'конференция', 'кулак', 'доска', 'ожидание', 'платье', 'смех', 'отказ', 'сбор', 'пенсия', 'буква', 'порог', 'автобус', 'воспитание', 'производитель', 'полоса', 'риск', 'пиво', 'корпус', 'штаб', 'кольцо', 'постель', 'выпуск', 'дворец', 'брак', 'прокурор', 'печать', 'окончание', 'автомат', 'тенденция', 'следователь', 'штат', 'куст', 'старуха', 'описание', 'психология', 'шутка', 'съезд', 'ставка', 'забота', 'величина', 'версия', 'мешок', 'конструкция', 'контакт', 'шанс', 'лодка', 'редактор', 'заказ', 'кофе', 'рубеж', 'статус', 'спорт', 'покой', 'кризис', 'взрыв', 'профессия', 'дым', 'металл', 'сапог', 'диван', 'интернет', 'почва', 'лед', 'подразделение', 'минимум', 'конь', 'дружба', 'вина', 'замок', 'мечта', 'сигнал', 'талант', 'мгновение', 'столик', 'затрата', 'золото', 'миг', 'плата', 'подъезд', 'масштаб', 'обсуждение', 'сделка', 'обязательство', 'расстояние', 'отдых', 'телевидение', 'тетя', 'яблоко', 'свидетель', 'монастырь', 'чтение', 'параметр', 'кампания', 'помощник', 'полк', 'мощность', 'сюжет', 'потолок', 'регистрация', 'майор', 'эксплуатация', 'озеро', 'новое', 'атмосфера', 'премия', 'совесть', 'предприниматель', 'мальчишка', 'дочка', 'приятель', 'начальство', 'препарат', 'село', 'обработка', 'танк', 'милиционер', 'ручка', 'возвращение', 'прокуратура', 'ворота', 'молоко', 'еда', 'сказка', 'краска', 'хвост', 'сигарета', 'введение', 'покупатель', 'поворот', 'москвич', 'ограничение', 'инвестиция', 'нация', 'набор', 'поселок', 'дыхание', 'адвокат', 'сумка', 'пресса', 'корреспондент', 'песок', 'удивление', 'потребитель', 'указание', 'изображение', 'счастье', 'мэр', 'согласие', 'действительность', 'планета', 'агентство', 'танец', 'библиотека', 'финансирование', 'объяснение', 'распределение', 'конституция', 'таблица', 'поэзия', 'термин', 'прибыль', 'стандарт', 'восторг', 'гибель', 'изделие', 'темп', 'вооружение', 'осуществление', 'уход', 'чемпионат', 'молитва', 'контракт', 'философия', 'горло', 'оборот', 'кость', 'ведомство', 'преимущество', 'мина', 'полномочие']
bot = telebot.TeleBot('*****')
counter = 0
@bot.message_handler(commands = ['start'])

def start(message):
    bot.reply_to(message, 'Напиши "!start game" ')
    # markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    # button1 = types.KeyboardButton("!rsdf")
    # button2 = types.KeyboardButton("!ldsf")
    # button3 = types.KeyboardButton("!start game")
    # bot.send_message(message.chat.id, "бури буриa", reply_markup=markup)
    # markup.add(button1, button2, button3)


@bot.message_handler(content_types='text')
def message_reply(message):
    global counter
    if message.text == "!start game":
        if counter == 0:
            counter += 1
            start_game(message)
        elif counter == 1:
            bot.reply_to(message, "игра уже запущена")
        print(letters)
    elif message.text == "!stop game" or message == "!stop game":
        if counter == 1:
            bot.reply_to(message, "игра окончена")
            counter = 0
        elif counter == 0:
            bot.reply_to(message, "игра не запущена")
    elif message.text == "!Avtoriza" or message.text == "!avtoriza" or message == "!Avtoriza" or message == "!avtoriza":
        bot.send_message(message.chat.id, "login ,reg, add")
        bot.register_next_step_handler(message, bazad)
        # bazad(message)
def start_game(message):
    global a
    global ans
    a = random.randint(0,999)
    global s
    s = " "
    for i in range(len(words[a])):
        s += "◻"
    for i in range(len(words[a])):
        ans += "◻"
    print(words[a])
    bot.send_message(message.chat.id, s)
    bot.register_next_step_handler(message, game)

def game(message):
    kolvo = 0
    global players
    global points
    global ans
    global nick
    global correctwords
    if message == "!stop game"or message.text == "!stop game" :
        message_reply(message)
        letters.clear()
        correctwords = []
        word, letter, ans, players, points = "", "", "", set(), dict()
    elif message.text == "!start game" or message == "!start game":
        message_reply(message)
    elif message.text[0] != "!":
        bot.register_next_step_handler(message, game)
    elif message.text == "!":
        bot.reply_to(message, "Букву надо написать")
        bot.register_next_step_handler(message, game)
    else:
        nick = message.from_user.username
        if nick not in players:
            players.add(nick)
            points[nick] = 0
        letter = str(message.text)
        word = str(words[a])
        print(letter)
        if (letter[-1].upper() in letters or letter[-1].lower() in letters) and len(letter) == 2 and letter[0] == "!":
            bot.reply_to(message, f"Буква '{letter[-1]}' уже называлась \n @{nick} - 2 очка")
            points[nick] -= 2
            bot.register_next_step_handler(message, game)
        else:
            letters.add(letter[-1])
            if (letter[1:] == word or letter[1:].lower() == word) and letter[0] == "!":
                lok = f"Правильно, @{nick} угадал слово и получает {3 * len(word) - len(correctwords)} баллов \n это было слово '{word.upper()}'"
                bot.reply_to(message, lok)
                points[nick] += 3 * (len(word) - len(correctwords))
                while len(correctwords) != len(word):
                    correctwords.append("1")
            elif letter[1:] != word and letter[0] == "!" and len(letter) > 2:
                bot.reply_to(message, "неПравильно, твои баллы онулированны")
                points[nick] -= 0
                bot.register_next_step_handler(message, game)
            elif (letter[-1].upper() in word or letter[-1].lower() in word) and len(letter) == 2 and letter[0] == "!":
                for i in range(len(ans)):
                    if letter[-1].upper() == word[i] or letter[-1].lower() == word[i]:
                        ans = ans[:i] + str(letter[-1].upper()) + ans[i + 1:]
                        correctwords.append(letter[-1])
                        kolvo += 1
                global qw
                # qw = 'Буква "' + str(letter[-1]).upper()  + '" есть в слове' +'\n' + str(ans)
                qw = f"Буква ' {(letter[-1]).upper()} ' есть в слове \n @{nick} + {3 * kolvo} очка \n {ans}"
                points[nick] += 3 * kolvo
                bot.reply_to(message, qw)
                bot.register_next_step_handler(message, game)

            elif letter[-1].upper() not in word and letter[-1].lower() not in word and len(letter) == 2 and letter[0] == "!":
                g = "Буквы " + '"' + str(letter[-1]).upper() + '"' + " нет в слове" + '\n' + str(ans)
                bot.reply_to(message, g)
                bot.register_next_step_handler(message, game)

            if len(correctwords) == len(word):
                max = points[nick]
                global winner
                winner = nick
                for key, value in points.items():
                    if value > max:
                        max = value
                        winner = key
                k = f"Вы угадали слово, это было слово: {(word).upper()} \n победитель : @{winner} с {points[winner]} баллами \n"
                for key, value in points.items():
                    k += "@"
                    k += str(key)
                    k += " : "
                    k += str(value)
                bot.send_message(message.chat.id, k)
                global counter
                counter = 0
                letters.clear()
                correctwords = []
                word, letter, ans, players, points = "", "", "", set(), dict()
                print(points)
            # message_reply(message)
        # bot.register_next_step_handler(message, game)

def bazad(message):
    global sheet
    sheet = book.active
    check_true = 1
    # bot.send_message(message.chat.id, "бури бури")
    if message.text.lower() == "!добавить" or message.text.lower() == "!add":
        if check_true == 1:
            bot.send_message(message.chat.id, "в какой отдел добавить? (write key)")
            bot.register_next_step_handler(message, addmes)
            # addmes(message)
        else:
            bot.send_message(message.chat.id, "войди в аккаунт")
    elif message.text.lower() == "!register" or message.text.lower() == "!reg" or message.text.lower() == "!sign in":
        for i in range(1, sheet.max_row + 1):
            if sheet["A" + str(i)].value == message.from_user.username:
                bot.send_message(message.chat.id, "вы уже зарегестрированны")
                bot.register_next_step_handler(message, message_reply)
                break
        else:
            bot.send_message(message.chat.id, "otprav login")
            bot.register_next_step_handler(message, addlogin)
    elif message.text.lower() == "!login" or message.text.lower() == "!log":
        bot.send_message(message.chat.id, "отправь свой логин")
        bot.register_next_step_handler(message, check_log)
    else:
        bot.register_next_step_handler(message, bazad)


def addlogin(message):
    global sheet
    global book
    sheet.cell(row=sheet.max_row + 1, column=1).value = message.from_user.username
    sheet.cell(row= sheet.max_row, column=2).value = message.text
    book.save("da.xlsx")
    # sheet_2 = book.worksheets[1]
    # sheet_2.cell(row=sheet.max_row, column=1).value = message.text
    # sheet = book.active
    # book.save("da.xlsx")
    bot.delete_message(message.chat.id, message.message_id)
    bot.send_message(message.chat.id, "логин успешно сохранен? теперь напиши пароль")
    bot.register_next_step_handler(message, addpassword)

def addpassword(message):
    global sheet
    sheet.cell(row=sheet.max_row, column=3).value = message.text
    bot.delete_message(message.chat.id, message.message_id)
    bot.send_message(message.chat.id, "пароль успешно сохранен?")
    sheet.cell(row=sheet.max_row, column=4).value = 1
    book.save("da.xlsx")
    # sheet_2 = book.worksheets[1]
    # sheet_2.cell(row=sheet.max_row, column=2).value = message.text
    # book.save("da.xlsx")
    bot.register_next_step_handler(message, message_reply)

def check_log(message):
    counter_chek_log = 0
    global sheet
    global book
    global row_login
    for i in range(1, sheet.max_row + 1):
        if sheet[i][0].value == message.from_user.username:
            if sheet[i][1].value == message.text[1:] or sheet[i][1].value == message.text:
                bot.delete_message(message.chat.id, message.message_id)
                bot.send_message(message.chat.id, "такой пользователь есть, отправь пароль")
                counter_chek_log += 1
                row_login = i
                bot.register_next_step_handler(message, check_pass)
                break
    if counter_chek_log == 0:
        bot.send_message(message.chat.id, "обманщик?")
        bot.register_next_step_handler(message, message_reply)
    # else:
    #     pass

def check_pass(message):
    global sheet
    global book
    global row_login
    if sheet[row_login][2].value == message.text[1:] or sheet[row_login][2].value == message.text:
        bot.delete_message(message.chat.id, message.message_id)
        bot.send_message(message.chat.id, "вход успешно выполнен?")
        sheet.cell(row=row_login, column=4).value = 1
        # sheet[row_login][3].value = 1
        book.save("da.xlsx")
        bot.register_next_step_handler(message, message_reply)
    else:
        bot.send_message(message.chat.id, "wrong pass, :(")
        bot.register_next_step_handler(message, message_reply)


def addmes(message):
    global counter_addmes
    global key_id
    global key_id2
    bot.send_message(message.chat.id, "успешно")
    # bot.send_message(message.chat.id, "в какой отдел добавить? (write key)")
    if counter_addmes == 0:
        # bot.send_message(message.chat.id, "в какой отдел добавить? (write key)")
        for i in range(1, sheet.max_row + 1):
            if sheet[i][0].value == message.from_user.username:
                global key_id
                key_id = i
                break
        # for i in range(sheet.max_column):
        #     if sheet[key_id][i].value == " ":
        #         print("asdadasd")
        #         global key_id2
        #         key_id2 = i
        #         break
        sheet.cell(row=key_id, column=sheet.max_column + 1).value = message.text
        print(key_id)
        book.save("da.xlsx")
        counter_addmes += 1
        bot.register_next_step_handler(message, addmes)
    elif counter_addmes == 1:
        print(message.text)
        sheet.cell(row=1, column=sheet.max_column + 1).value = message.text
        book.save("da.xlsx")
        counter_addmes += 1
        bot.register_next_step_handler(message, addmes)
    elif counter_addmes == 2:
        counter_addmes = 0
        print(message.text)


bot.polling(none_stop=True)
